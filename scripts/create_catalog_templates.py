from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
SCHEMA_TEMPLATE_DIR = ROOT / "schemas" / "templates"
PUBLIC_TEMPLATE_DIR = ROOT / "public" / "catalog-templates"

COLUMNS = [
    ("external_id", "필수", "판매처에서 유지되는 상품 ID", "EXAMPLE-001"),
    ("name", "필수", "상품 표시명", "입력 예시 상품"),
    ("brand", "필수", "브랜드명. 비우면 전용 config 기본값 사용", "한샘"),
    ("sku", "선택", "판매처 SKU/품번", "EXAMPLE-001"),
    ("category", "필수", "HomePlan 표준 taxonomy", "kitchen.base-cabinet"),
    ("width", "필수", "가로 치수", "800"),
    ("depth", "필수", "깊이 치수", "600"),
    ("height", "필수", "높이 치수", "850"),
    ("unit", "필수", "치수 단위: mm, cm, m", "mm"),
    ("source_url", "필수", "근거 상품 페이지 HTTPS URL", "https://example.com/product/1"),
    ("retrieved_at", "필수", "근거 확인 ISO 일시", "2026-08-31T00:00:00+09:00"),
    ("mount", "필수", "floor, wall, wall-mount, ceiling, surface", "floor"),
    ("price_amount", "선택", "정수 원화 가격", "490000"),
    ("price_currency", "가격 입력 시 필수", "v1은 KRW", "KRW"),
    ("price_checked_at", "가격 입력 시 필수", "가격 확인일 YYYY-MM-DD", "2026-08-31"),
    ("price_basis", "가격 입력 시 필수", "가격에 포함되는 구성 기준", "본체 1개"),
    ("included", "선택", "가격 포함 항목. 여러 값은 | 구분", "본체|기본 하드웨어"),
    ("excluded", "선택", "가격 제외 항목. 여러 값은 | 구분", "배송|시공"),
    ("materials", "선택", "소재. 여러 값은 | 구분", "PB|MDF"),
    ("tags", "선택", "검색 태그. 여러 값은 | 구분", "주방|하부장"),
    ("images", "선택", "HTTPS 이미지 URL. 여러 값은 | 구분", "https://example.com/image.jpg"),
    ("colorways", "선택", "#RRGGBB 색상. 여러 값은 | 구분", "#F2F0EA|#3A3A38"),
    ("provides", "선택", "제공 capability. 여러 값은 | 구분", "support.base-cabinet|support.countertop"),
    ("requires_all", "선택", "모두 필요한 capability. 여러 값은 | 구분", ""),
    ("requires_any", "선택", "하나 이상 필요한 capability. 여러 값은 | 구분", ""),
    ("dependency_scope", "의존성 입력 시 선택", "support-chain 또는 project", "support-chain"),
    ("surface_supported_by", "surface 배치 시 필수", "받침 capability. 여러 값은 | 구분", ""),
    ("surface_anchor", "선택", "rear 또는 center", "rear"),
    ("shape_hint", "선택", "지원 렌더 shape. 미지원 값은 box로 대체", "sinkLower"),
    ("snap_to_wall", "선택", "true/false", "true"),
    ("default_elevation", "선택", "기본 설치 높이(mm)", ""),
]

CATEGORIES = [
    "seating.sofa",
    "seating.chair",
    "table.coffee",
    "table.dining",
    "bed.frame",
    "storage.wardrobe",
    "storage.bookcase",
    "storage.built-in",
    "kitchen.base-cabinet",
    "kitchen.wall-cabinet",
    "kitchen.sink",
    "kitchen.faucet",
    "kitchen.appliance",
    "appliance",
    "lighting",
    "bathroom",
    "curtain",
    "finish.wallcovering",
    "finish.flooring",
]

BRANDS = {
    "hanssem": {
        "provider": "Hanssem",
        "brand": "한샘",
        "catalog_id": "hanssem-ko",
        "sample": {
            "external_id": "EXAMPLE-HS-KITCHEN-001",
            "name": "입력 예시 - 한샘 주방 하부장",
            "sku": "EXAMPLE-HS-KITCHEN-001",
            "category": "kitchen.base-cabinet",
            "width": 800,
            "depth": 600,
            "height": 850,
            "source_url": "https://store.hanssem.com/",
            "mount": "floor",
            "price_amount": 490000,
            "price_basis": "예시 본체 1개",
            "materials": "PB|MDF",
            "tags": "입력예시|주방|하부장",
            "provides": "support.base-cabinet|support.countertop",
            "shape_hint": "sinkLower",
            "snap_to_wall": True,
        },
    },
    "livart": {
        "provider": "Hyundai Livart",
        "brand": "리바트",
        "catalog_id": "hyundai-livart-ko",
        "sample": {
            "external_id": "EXAMPLE-LV-SOFA-001",
            "name": "입력 예시 - 리바트 4인 소파",
            "sku": "EXAMPLE-LV-SOFA-001",
            "category": "seating.sofa",
            "width": 3000,
            "depth": 950,
            "height": 870,
            "source_url": "https://www.hyundailivart.co.kr/",
            "mount": "floor",
            "price_amount": 2237000,
            "price_basis": "예시 본체 1개",
            "materials": "천연가죽|PVC",
            "tags": "입력예시|거실|4인용",
            "provides": "seating.sofa",
            "shape_hint": "sofa3",
            "snap_to_wall": False,
        },
    },
}


def complete_sample(spec: dict) -> dict:
    sample = dict(spec["sample"])
    sample.update(
        {
            "brand": spec["brand"],
            "unit": "mm",
            "retrieved_at": "2026-08-31T00:00:00+09:00",
            "price_currency": "KRW",
            "price_checked_at": "2026-08-31",
            "dependency_scope": "support-chain",
            "surface_anchor": "rear",
        }
    )
    return sample


def create_workbook(slug: str, spec: dict) -> Path:
    workbook = Workbook()
    products = workbook.active
    products.title = "products"
    products.sheet_view.showGridLines = False
    products.freeze_panes = "A2"

    headers = [column[0] for column in COLUMNS]
    sample = complete_sample(spec)
    products.append(headers)
    products.append([sample.get(header, "") for header in headers])
    products.auto_filter.ref = f"A1:{get_column_letter(len(headers))}2"

    header_fill = PatternFill("solid", fgColor="173F47")
    input_fill = PatternFill("solid", fgColor="EAF4F7")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    for index, (name, requirement, description, example) in enumerate(COLUMNS, start=1):
        cell = products.cell(1, index)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = Comment(f"{requirement}\n{description}\n예: {example}", "HomePlan")
        products.cell(2, index).fill = input_fill
        products.cell(2, index).font = Font(color="1F4E78")
        width = max(13, min(34, len(name) + 6))
        products.column_dimensions[get_column_letter(index)].width = width
    products.row_dimensions[1].height = 30

    header_index = {header: index + 1 for index, header in enumerate(headers)}
    validations = {
        "category": CATEGORIES,
        "unit": ["mm", "cm", "m"],
        "mount": ["floor", "wall", "wall-mount", "ceiling", "surface"],
        "dependency_scope": ["support-chain", "project"],
        "surface_anchor": ["rear", "center"],
        "snap_to_wall": ["true", "false"],
    }
    lists = workbook.create_sheet("lists")
    lists.sheet_state = "hidden"
    for validation_index, (header, choices) in enumerate(validations.items(), start=1):
        for row_index, choice in enumerate(choices, start=1):
            lists.cell(row_index, validation_index, choice)
        validation_column = get_column_letter(validation_index)
        formula = f"=lists!${validation_column}$1:${validation_column}${len(choices)}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "목록에 있는 표준값을 선택해 주세요."
        validation.errorTitle = "지원하지 않는 값"
        validation.prompt = "HomePlan 표준값을 선택하세요."
        validation.promptTitle = header
        validation.showErrorMessage = True
        validation.showInputMessage = True
        products.add_data_validation(validation)
        letter = get_column_letter(header_index[header])
        validation.add(f"{letter}2:{letter}1001")

    required_headers = [
        "external_id",
        "name",
        "category",
        "width",
        "depth",
        "height",
        "source_url",
        "retrieved_at",
    ]
    for header in required_headers:
        letter = get_column_letter(header_index[header])
        products.conditional_formatting.add(
            f"{letter}2:{letter}1001",
            FormulaRule(formula=[f'LEN({letter}2)=0'], fill=warning_fill),
        )

    guide = workbook.create_sheet("guide")
    guide.sheet_view.showGridLines = False
    guide.append([f"{spec['brand']} 카탈로그 bridge", "HomePlan Catalog Protocol 1.0"])
    guide.append(["사용 순서", "products 시트의 예시 행을 복제/수정 → CSV 또는 XLSX 변환 → 프로토콜 검증"])
    guide.append(["주의", "예시 행은 실상품 데이터가 아닙니다. source_url과 확인일을 실제 근거로 교체하세요."])
    guide.append([])
    guide.append(["컬럼", "필수 여부", "설명", "예시"])
    for row in COLUMNS:
        guide.append(list(row))
    guide.freeze_panes = "A6"
    guide.auto_filter.ref = f"A5:D{5 + len(COLUMNS)}"
    guide.column_dimensions["A"].width = 25
    guide.column_dimensions["B"].width = 20
    guide.column_dimensions["C"].width = 65
    guide.column_dimensions["D"].width = 48
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=13)
    for cell in guide[5]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    for row in guide.iter_rows(min_row=1, max_row=guide.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.properties.creator = "HomePlan 3D"
    workbook.properties.title = f"{spec['brand']} HomePlan 카탈로그 템플릿"
    workbook.properties.description = "HomePlan Catalog Protocol 1.0 CSV/XLSX bridge"
    path = PUBLIC_TEMPLATE_DIR / f"{slug}-catalog-template.xlsx"
    workbook.save(path)
    return path


def create_csv(slug: str, spec: dict) -> Path:
    path = PUBLIC_TEMPLATE_DIR / f"{slug}-catalog-template.csv"
    headers = [column[0] for column in COLUMNS]
    sample = complete_sample(spec)
    with path.open("w", newline="", encoding="utf-8-sig") as stream:
        writer = csv.DictWriter(stream, fieldnames=headers)
        writer.writeheader()
        writer.writerow(sample)
    return path


def create_config(slug: str, spec: dict) -> Path:
    path = SCHEMA_TEMPLATE_DIR / f"{slug}-sheet.config.json"
    config = {
        "catalog": {
            "id": spec["catalog_id"],
            "provider": spec["provider"],
            "locale": "ko-KR",
            "generatedAt": "2026-08-31T00:00:00.000Z",
        },
        "defaults": {"brand": spec["brand"], "unit": "mm", "mount": "floor"},
    }
    path.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def main() -> int:
    SCHEMA_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    created = []
    for slug, spec in BRANDS.items():
        created.extend(
            [create_workbook(slug, spec), create_csv(slug, spec), create_config(slug, spec)]
        )
    for path in created:
        print(path.relative_to(ROOT))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
