from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def serializable(value):
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return value


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    parser.add_argument("--sheet", default="products")
    args = parser.parse_args()
    path = Path(args.input).resolve()
    formulas = load_workbook(path, read_only=True, data_only=False)
    values = load_workbook(path, read_only=True, data_only=True)
    if args.sheet not in formulas.sheetnames:
        raise SystemExit(f"sheet-not-found:{args.sheet}")
    formula_sheet = formulas[args.sheet]
    value_sheet = values[args.sheet]
    formula_rows = formula_sheet.iter_rows()
    value_rows = value_sheet.iter_rows()
    headers = [str(cell.value or "").strip() for cell in next(formula_rows)]
    next(value_rows)
    rows = []
    for formula_row, value_row in zip(formula_rows, value_rows):
        if not any(cell.value not in (None, "") for cell in formula_row):
            continue
        output = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            formula_value = formula_row[index].value
            cached_value = value_row[index].value
            if isinstance(formula_value, str) and formula_value.startswith("="):
                if cached_value is None:
                    raise SystemExit(
                        f"formula-cache-missing:{args.sheet}!{formula_row[index].coordinate}"
                    )
                output[header] = serializable(cached_value)
            else:
                output[header] = serializable(formula_value)
        rows.append(output)
    json.dump(rows, sys.stdout, ensure_ascii=False)
    formulas.close()
    values.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
